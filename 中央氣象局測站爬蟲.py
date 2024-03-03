from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
import openpyxl
import time
from datetime import datetime

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--headless')
chrome_options.add_argument('--disable-gpu')

driver = webdriver.Chrome(service=ChromeService(), options=chrome_options)
driver.get('https://codis.cwa.gov.tw/StationData')
driver.set_window_size(1660, 945)
time.sleep(1)

#改成清單模式
driver.find_element(By.CSS_SELECTOR, '#switch_display > button:nth-child(2)').click()

#挑選不同類型的測站
#driver.find_element(By.CSS_SELECTOR, '#cwb').click() #署屬有人站
#driver.find_element(By.CSS_SELECTOR, '#auto_C1').click() #自動雨量站
#driver.find_element(By.CSS_SELECTOR, '#auto_C0').click() #自動氣象站
#driver.find_element(By.CSS_SELECTOR, '#agr').click() #農業站

workbook = openpyxl.Workbook()
sheet = workbook.worksheets[0]

#選擇站點
times = driver.find_element(By.CSS_SELECTOR, '#station_count').text
for nowtimes in range(1, int(times)+1):
    name = driver.find_element(By.CSS_SELECTOR, '#station_table > table > tbody > tr:nth-child({}) > td:nth-child(1) > div'.format(nowtimes)).text
    if('雷達站' in name):
        print("雷達站跳過")
        continue
    print("目前測站{}，還有{}個測站".format(name, str(int(times)-nowtimes)))
    driver.find_element(By.CSS_SELECTOR, '#station_table > table > tbody > tr:nth-child({}) > td:nth-child(10) > div'.format(nowtimes)).click()
    time.sleep(2)
    
    
    #提取站點資訊
    place = driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div.lightbox-tool-station > div.lightbox-tool-station-info > div.lightbox-tool-station-subinfo > div:nth-child(1)').text.split('：')[1]
    longitude = driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div.lightbox-tool-station > div.lightbox-tool-station-info > div.lightbox-tool-station-subinfo > div:nth-child(2)').text.split('：')[1].split(',')[0][:-2]
    latitude = driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div.lightbox-tool-station > div.lightbox-tool-station-info > div.lightbox-tool-station-subinfo > div:nth-child(2)').text.split('：')[1].split(',')[1][:-2]
    altitude = driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div.lightbox-tool-station > div.lightbox-tool-station-info > div.lightbox-tool-station-subinfo > div:nth-child(3)').text.split('：')[1]
    date = driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div.lightbox-tool-station > div.lightbox-tool-station-info > div.lightbox-tool-station-subinfo > div:nth-child(4)').text[5:15]
    date = list(map(int, date.split('-')))
    if(date[0] < 1990):
        date = [1990, 1, 1]
    #修改觀測年月日
    driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > label > div > div.vdatetime > input').click()
    time.sleep(1)

    #年
    nyear = driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > label > div > div.vdatetime > div > div.vdatetime-popup > div.vdatetime-popup__header > div.vdatetime-popup__year').text
    driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > label > div > div.vdatetime > div > div.vdatetime-popup > div.vdatetime-popup__header > div.vdatetime-popup__year').click()
    driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > label > div > div.vdatetime > div > div.vdatetime-popup > div.vdatetime-popup__body > div > div > div:nth-child({})'.format(100-(int(nyear)-date[0]-1))).click()

    now = driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > label > div > div.vdatetime > div > div.vdatetime-popup > div.vdatetime-popup__header > div.vdatetime-popup__date').text
    #月
    nmin = now.split('月')[0]
    driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > label > div > div.vdatetime > div > div.vdatetime-popup > div.vdatetime-popup__header > div.vdatetime-popup__date').click()
    if(date[1] == nmin):
        driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > label > div > div.vdatetime > div > div.vdatetime-popup > div.vdatetime-popup__body > div > div > div.vdatetime-month-picker__item.vdatetime-month-picker__item--selected').click()
    else:
        driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > label > div > div.vdatetime > div > div.vdatetime-popup > div.vdatetime-popup__body > div > div > div:nth-child({})'.format(date[1])).click()

    #日
    nday = now.split('月')[1].replace('日', '')
    if(date[2] == nday):
        driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > label > div > div.vdatetime > div > div.vdatetime-popup > div.vdatetime-popup__body > div > div.vdatetime-calendar__month > div.vdatetime-calendar__month__day.vdatetime-calendar__month__day--selected').click()
    else:
        tmp = driver.find_elements(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > label > div > div.vdatetime > div > div.vdatetime-popup > div.vdatetime-popup__body > div > div.vdatetime-calendar__month > div')
        for i in tmp:
            if(i.text == str(date[2])):
                i.click()
                break

    driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > label > div > div.vdatetime > div > div.vdatetime-popup > div.vdatetime-popup__actions > div.vdatetime-popup__actions__button.vdatetime-popup__actions__button--confirm').click()
    time.sleep(1)

    #計算該觀測站共有幾天資料須爬取
    if(date == [2023, 12, 31]):
        daynum = 1
    else:
        daynum = datetime(2023, 12, 31) - datetime(date[0], date[1], date[2])
        daynum = int(str(daynum).split(' ')[0])+1
    print(daynum)

    #將資料輸入進excel
    sheet.cell(row = 1, column = 1).value = "站名"
    sheet.cell(row = 1, column = 2).value = "縣市"
    sheet.cell(row = 1, column = 3).value = "經度"
    sheet.cell(row = 1, column = 4).value = "緯度"
    sheet.cell(row = 1, column = 5).value = "海拔高度"
    sheet.cell(row = 1, column = 6).value = "觀測日期"
    sheet.cell(row = 2, column = 1).value = "Name"
    sheet.cell(row = 2, column = 2).value = "Place"
    sheet.cell(row = 2, column = 3).value = "Longitude"
    sheet.cell(row = 2, column = 4).value = "Latitude"
    sheet.cell(row = 2, column = 5).value = "Altitude"
    sheet.cell(row = 2, column = 6).value = "Obsdate"
    
    chlist = driver.find_elements(By.CSS_SELECTOR, '#report_date > table > thead > tr:nth-child(3) > th')
    enlist = driver.find_elements(By.CSS_SELECTOR, '#report_date > table > thead > tr:nth-child(4) > th')
    for i in range(len(chlist)):
        sheet.cell(row = 1, column = i + 7).value = chlist[i].text
        sheet.cell(row = 2, column = i + 7).value = enlist[i].text

    nrow = 3
    for _ in range(daynum):
        print("目前測站還有{}天".format(str(daynum-_-1)))
        for num in range(1, 25):
            sheet.cell(row = nrow, column = 1).value = name
            sheet.cell(row = nrow, column = 2).value = place
            sheet.cell(row = nrow, column = 3).value = longitude
            sheet.cell(row = nrow, column = 4).value = latitude
            sheet.cell(row = nrow, column = 5).value = altitude
            sheet.cell(row = nrow, column = 6).value = '/'.join(list(map(str, date)))
            for i in range(len(chlist)):
                sheet.cell(row = nrow, column = i + 7).value = driver.find_element(By.CSS_SELECTOR, '#report_date > table > tbody > tr:nth-child({}) > td:nth-child({}) > div'.format(str(num), str(i+1))).text
            nrow += 1
        driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > div > section > div:nth-child(5) > div.lightbox-tool-type-ctrl > div.lightbox-tool-type-ctrl-form > label > div > div:nth-child(5)').click()
        time.sleep(1)

        if(_%7 == 0):
            workbook.save('{}.xlsx'.format(name))
    workbook.save('{}.xlsx'.format(name))
    workbook.close()
    driver.find_element(By.CSS_SELECTOR, '#main_content > section.lightbox-tool > div > header > div.lightbox-tool-close').click()
    time.sleep(1)