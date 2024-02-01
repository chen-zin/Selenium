import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
import time
import warnings #忽略警告訊息
import traceback #在except看錯誤資訊
import re #用正則表達式來替換字

warnings.simplefilter('ignore', ResourceWarning)
strinfo = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

#用來轉換經緯度格式的
def DMS2DD(old):
    #ex：24°57'20"N
    D, old = old.split('°')
    M, old = old.split("'")
    S, old = old.split('"')
    
    D = int(D)
    M = int(M)
    S = float(S) if ('.' in S) else int(S)

    return round(D + M/60 + S/3600, 7)

'''
設定參數
maxnum：網站總筆數，在租屋見證網頁右下角查看或是網址totalRows後的數字
firstpage：要從網站哪一頁開始爬
year：網站內沒有年份資訊，所以自己添加，如果從第一頁開始就是2024，不是的話看之前最後一筆數據是哪一年決定
'''
#租屋見證網站：https://www.591.com.tw/witness.html?list=1
maxnum = 389
firstpage = 1 - 1
year = 2017

workbook = openpyxl.load_workbook('成交data.xlsx')
sheet = workbook.worksheets[0]
row = sheet.max_row
row_tmp = row
flag = True
last_min = 0

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--headless') # 啟動無頭模式
chrome_options.add_argument('--disable-gpu') # windows必須加入此行

#執行中偶爾會有網頁加載太久結果報錯，增加判斷，如果報錯時不是在最後一頁就關閉driver然後重開一個
while(firstpage < maxnum//10):
    try:
        driver = webdriver.Chrome(service=ChromeService(), options=chrome_options)
        driver.set_page_load_timeout(90)

        for page in range(firstpage, maxnum//10 +1):
            
            firstpage = page
            driver.get("https://www.591.com.tw/index.php?firstRow={}&totalRows={}&list=1&module=guestBook&action=witness".format(str(page*10), str(maxnum)))
            time.sleep(1)

            href = []
            ID = []
            date = []
            times = []
            human = []

            #在出租見證網頁抓取連結與所需資訊
            for i in range(5,15):

                #單個出租資訊的連結
                tmp = driver.find_element(By.CSS_SELECTOR,"#Content > div.KfColRight > div:nth-child({}) > div.Mbottom > div > div.CTop > a:nth-child(5)".format(str(i))).get_attribute('href')
                href.append(tmp)
                ID.append(tmp.split('.')[-2].split('-')[-1])

                #上架及成交時間，補上年份，特別注意1月和12月
                tmp = driver.find_element(By.CSS_SELECTOR,"#Content > div.KfColRight > div:nth-child({}) > div.Mtop > div.About".format(str(i))).text
                tt = False
                if(flag and int(tmp[20:22]) == 12 and last_min != 11):
                    year -= 1
                    flag = False
                #偶爾在12月的數據裡會混入1月的資料，為了防止年份亂掉所以設計的
                elif(not(flag) and int(tmp[20:22]) == 1):
                    tt = True
                elif(not(flag) and int(tmp[20:22]) == 11):
                    flag = True
                else:
                    pass

                last_min = int(tmp[20:22])
                
                if(tt):
                    #偶爾在12月的數據裡會混入1月的資料，為了防止年份亂掉所以設計的
                    tmp = [ str(year if int(tmp[3:5]) > int(tmp[20:22]) else year+1)  + "/" + tmp[3:5] + "/" + tmp[6:8],\
                            str(year) + "/" + tmp[20:22] + "/" + tmp[23:25]]
                else:
                    tmp = [ str(year-1 if int(tmp[3:5]) > int(tmp[20:22]) else year)  + "/" + tmp[3:5] + "/" + tmp[6:8],\
                            str(year) + "/" + tmp[20:22] + "/" + tmp[23:25]]
                date.append(tmp)

                #成交天數，有兩種可能所以要用 try 防止 error
                try:
                    tmp = driver.find_element(By.CSS_SELECTOR,"#Content > div.KfColRight > div:nth-child({}) > div.Mbottom > div > div.CTop > span.DealEnd".format(str(i))).text
                except:
                    tmp = driver.find_element(By.CSS_SELECTOR,"#Content > div.KfColRight > div:nth-child({}) > div.Mbottom > div > div.CTop > span.DealEnd2".format(str(i))).text
                times.append(tmp.split('天')[0])

                #招租人
                tmp = driver.find_element(By.CSS_SELECTOR,"#Content > div.KfColRight > div:nth-child({}) > div.Mtop > div.Nickname".format(str(i))).text
                human.append(tmp)
                
            
            #依序打開連結抓取所需資訊
            for i in range(len(href)):
                try:

                    data = ['' for _ in range(43)]
                    driver.get(href[i])
                    time.sleep(2)

                    #租屋類型，判斷是不是需要的，順便篩掉已經下架的
                    try:
                        Type = driver.find_element(By.CSS_SELECTOR,"#app > div.bread-crumb-skeleton > section > div.bread-crumb > ul > li:nth-child(3) > a").text
                    except:
                        continue
                    if(not(Type == '整層住家' or Type == '獨立套房' or Type == '分租套房' or Type == '雅房' or Type == '其他')):
                        continue

                    print("第",row,"筆資料")
                    print(href[i])
                    print()

                    data[0] = str(ID[i])
                    data[5] = str(human[i])
                    data[7] = str(Type)
                    data[38] = str(date[i][0])
                    data[39] = str(times[i])
                    data[40] = str(date[i][1])

                    #名字，可能出現excel不要的字符，要先濾掉
                    tmp = driver.find_element(By.CSS_SELECTOR,"#houseInfo > div.house-title > h1").text
                    try:
                        tt = driver.find_element(By.CSS_SELECTOR,"#houseInfo > div.house-title > h1 > span").text
                        data[1] = tmp[:-1*len(tt)]
                    except:
                        data[1] = tmp
                    data[1] = strinfo.sub('', data[1])
                    
                    #縣市
                    tmp = driver.find_element(By.CSS_SELECTOR,"#app > div.bread-crumb-skeleton > section > div.bread-crumb > ul > li:nth-child(1) > a").text
                    data[2] = tmp

                    #區域
                    tmp = driver.find_element(By.CSS_SELECTOR,"#app > div.bread-crumb-skeleton > section > div.bread-crumb > ul > li:nth-child(2) > a").text
                    data[3] = tmp

                    #詳細地址
                    tmp = driver.find_element(By.CSS_SELECTOR,"#positionRound > div.address.ellipsis > p:nth-child(1) > span.load-map").text
                    data[4] = tmp[len(data[3]):]

                    #租金
                    tmp = driver.find_element(By.CSS_SELECTOR,"#houseInfo > div.house-price > span > b").text
                    data[6] = tmp

                    #房廳衛
                    tmp = driver.find_element(By.CSS_SELECTOR,"#houseInfo > div.house-pattern > span:nth-child(1)").text
                    data[8] = tmp

                    #坪數
                    tmp = driver.find_element(By.CSS_SELECTOR,"#houseInfo > div.house-pattern > span:nth-child(3)").text
                    data[9] = tmp

                    #樓層
                    tmp = driver.find_element(By.CSS_SELECTOR,"#houseInfo > div.house-pattern > span:nth-child(5)").text
                    data[10] = tmp
                    
                    #租屋型態
                    tmp = driver.find_element(By.CSS_SELECTOR,"#houseInfo > div.house-pattern > span:nth-child(7)").text
                    data[11] = tmp
                    
                    #押金
                    tmp = driver.find_element(By.CSS_SELECTOR,"#houseInfo > div.house-price").text
                    tt = driver.find_element(By.CSS_SELECTOR,"#houseInfo > div.house-price > span").text
                    data[29] = tmp[len(tt):]
                        
                    #規則，可能沒有
                    try:
                        tmp = driver.find_element(By.CSS_SELECTOR,"#service > div.service-rule > div > span").text
                    except:
                        tmp = '無'
                    data[12] = tmp
                    
                    #租金含，可能沒有
                    tmp = driver.find_element(By.CSS_SELECTOR,"#houseDetail > div.main-info-list > div.main-info-left > div.content > div:nth-child(1) > div.name > span").text
                    tt = driver.find_element(By.CSS_SELECTOR,"#houseDetail > div.main-info-list > div.main-info-left > div.content > div:nth-child(1) > div.text").text
                    if(tmp != '租金含'):
                        data[28] = '無'
                    else:
                        data[28] = tt
                    
                    #各種提供設備，共有15項，一個一個確定 class 後有沒有 del 來判斷有沒有這個設施，陽台要特別看數量、車位要看是哪個類型的
                    tmp = driver.find_element(By.CSS_SELECTOR,"#houseInfo > div.house-pattern > span:nth-child(5)").text
                    data[10] = tmp

                    for num in range(1,16):
                        tmp = driver.find_element(By.CSS_SELECTOR,"#service > div.service-list-box > div:nth-child({})".format(num)).get_attribute('class')
                        if(tmp[-3:] == 'del'):
                            data[num + 12] = '0'
                        elif(num == 13):
                            tmp = driver.find_element(By.CSS_SELECTOR,"#service > div.service-list-box > div:nth-child({}) > div".format(num)).text
                            data[num + 12] = tmp[:-2] if tmp[:-2] != '' else '1'
                            
                        elif(num == 15):
                            tmp = driver.find_element(By.CSS_SELECTOR,"#service > div.service-list-box > div:nth-child({}) > div".format(num)).text
                            data[num + 12] = tmp
                        else:
                            data[num + 12] = '1'
                    
                    fa = True
                    #打開地圖，位置與周邊 > 交通，有可能沒有地圖，如果沒有下述資訊都沒有，設計一個旗標檢測
                    try:
                        driver.find_element(By.CSS_SELECTOR,"#positionRound > div.surround-list > div:nth-child(1) > a").click()
                    except:
                        fa = False
                    
                    if(not(fa)):
                        data[30] = '無'
                        data[31] = '無'
                        data[32] = '無'
                        data[33] = '無'
                        data[34] = '無'
                        data[35] = '無'
                        data[36] = '無'
                        data[37] = '無'
                        data[41] = '無'
                        data[42] = '無'

                    else:
                        #經緯度
                        tmp = driver.find_element(By.CSS_SELECTOR,"#surround-map > div.jump-google-map > div.lat-lng").text
                        tmp_N, tmp_E = map(DMS2DD, tmp.split(' '))
                        
                        data[41] = tmp_E
                        data[42] = tmp_N

                        #最近交通，公車和捷運距離比較，一樣距離的話選捷運站
                        try:
                            tmp1 = driver.find_element(By.CSS_SELECTOR,"#surround-map > div.result-list > div.result-list-item.active > div > dl:nth-child(1) > dd:nth-child(2) > span.distance").text[1:-2]
                            tmp1 = int(tmp1)
                        except:
                            tmp1 = 99999999
                        try:
                            tmp2 = driver.find_element(By.CSS_SELECTOR,"#surround-map > div.result-list > div.result-list-item.active > div > dl:nth-child(2) > dd:nth-child(2) > span.distance").text[1:-2]
                            tmp2 = int(tmp2)
                        except:
                            tmp2 = 99999999
                        
                        if(tmp1 == 99999999 and tmp2 == 99999999):
                            data[30] = '無'
                            data[31] = '無'
                        else:
                            data[30] = tmp1 if tmp1 <= tmp2 else tmp2
                            data[31] = '捷運' if tmp1 <= tmp2 else '公車'
                        
                        #換成顯示 教育 頁面 > 國小
                        tmp = driver.find_element(By.CSS_SELECTOR,"#surround-map > div.result-list > ul > li.education").click()
                        
                        #最近國小，有可能沒有
                        try:
                            tmp = driver.find_element(By.CSS_SELECTOR,"#surround-map > div.result-list > div.result-list-item.active > div > dl.active > dd:nth-child(1) > span.distance").text[1:-2]
                            name = driver.find_element(By.CSS_SELECTOR,"#surround-map > div.result-list > div.result-list-item.active > div > dl.active > dd:nth-child(1) > span.name").text
                            data[33] = tmp
                            data[32] = name
                        except:
                            data[33] = '無'
                            data[32] = '無'

                        #換成顯示 教育 頁面 > 國中
                        tmp = driver.find_element(By.CSS_SELECTOR,"#surround-map > div.result-list > div.result-list-item.active > div > dl.live-type > dt:nth-child(2)").click()
                        
                        #最近國中，有可能沒有
                        try:
                            tmp = driver.find_element(By.CSS_SELECTOR,"#surround-map > div.result-list > div.result-list-item.active > div > dl.active > dd:nth-child(1) > span.distance").text[1:-2]
                            name = driver.find_element(By.CSS_SELECTOR,"#surround-map > div.result-list > div.result-list-item.active > div > dl.active > dd:nth-child(1) > span.name").text
                            data[35] = tmp
                            data[34] = name
                        except:
                            data[35] = '無'
                            data[34] = '無'
                        
                        #換成顯示 教育 頁面 > 國中
                        tmp = driver.find_element(By.CSS_SELECTOR,"#surround-map > div.result-list > div.result-list-item.active > div > dl.live-type > dt:nth-child(3)").click()

                        #最近大學，有可能沒有
                        try:
                            tmp = driver.find_element(By.CSS_SELECTOR,"#surround-map > div.result-list > div.result-list-item.active > div > dl.active > dd:nth-child(1) > span.distance").text[1:-2]
                            name = driver.find_element(By.CSS_SELECTOR,"#surround-map > div.result-list > div.result-list-item.active > div > dl.active > dd:nth-child(1) > span.name").text
                            data[37] = tmp
                            data[36] = name
                        except:
                            data[37] = '無'
                            data[36] = '無'
                except:
                    continue

                row += 1
                for col in range(len(data)):
                    sheet.cell(row = row, column = col + 1).value = data[col]
                
            #每一頁資訊寫入完存檔一次，順便列印目前總筆數、哪一頁、年份，以利重新啟動時設定
            workbook.save('成交data.xlsx')
            row_tmp = row
            print(row-1,"筆資料")
            print('目前年月：', date[-1][1][:7])
            print('下一頁', page+2, '頁', end='\n\n')
    
    except Exception as e:
        # 打印完整的異常信息
        traceback.print_exc()
        driver.quit()
        row = row_tmp
        print()
        print('重開driver，目前在第', firstpage+1, '頁')

driver.quit()
workbook.close()